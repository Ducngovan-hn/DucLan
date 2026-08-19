#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
dh.py — CÔNG CỤ ĐƠN HÀNG XƯỞNG BO ĐỨC LAN

Làm hai việc trong quy trình 7 bước của anh Đức:
  · Bước 3: ảnh cú pháp đặt bo  ->  Tong-hop-don-hang-Xbo-Duc-Lan-2026.xlsx
  · Bước 6: tin nhắn nhóm Zalo "Trả hàng"  ->  Xưởng Dệt Bo Đức Lan 2026.xlsx

Mọi việc tính toán, đối chiếu, chống trùng, ghi Excel đều làm ở đây (không tốn token).
LLM chỉ làm đúng một việc: đọc chữ trên ảnh / trên tin nhắn rồi xuất JSON.

Các lệnh:
  python tools/dh.py xay-tu-dien          Dựng lại tools/tu-dien.json
  python tools/dh.py ma-det "42x7" "kẻ"   Thử quy tắc đuôi mã
  python tools/dh.py anh-moi              Liệt kê + thu nhỏ ảnh chưa xử lý
  python tools/dh.py don-them --json F    Ghi đơn hàng vào Excel đơn hàng
  python tools/dh.py tra-them --json F    Ghi trả hàng vào sổ thật
  python tools/dh.py soat                 Rà soát sức khoẻ dữ liệu

Thêm --thu vào don-them/tra-them để CHẠY KHÔ (không ghi gì).
"""
from __future__ import annotations

import argparse
import copy as _copy
import datetime as dt
import difflib
import hashlib
import json
import re
import shutil
import sys
import unicodedata
from pathlib import Path

# ----- buộc stdout ra UTF-8 để tiếng Việt không lỗi trên Windows -----
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    sys.exit("Thiếu openpyxl. Chạy: pip install openpyxl")

# ============================ ĐƯỜNG DẪN ============================
VAULT = Path(__file__).resolve().parent.parent
XUONG = VAULT / "raw" / "xưởng  bo Đức Lan"   # LƯU Ý: hai dấu cách giữa "xưởng" và "bo"
ANH_DIR = XUONG / "Đơn hàng"
THIETKE_DIR = XUONG / "Thiết kế"

import os

# Cho phép trỏ sang file khác khi kiểm thử, để không ghi bẩn sổ thật:
#   set DH_XL_DON=...\ban-thu.xlsx
XL_DON = Path(os.environ.get("DH_XL_DON") or VAULT / "production" / "Tong-hop-don-hang-Xbo-Duc-Lan-2026.xlsx")
XL_SO = Path(os.environ.get("DH_XL_SO") or VAULT / "production" / "Xưởng Dệt Bo Đức Lan 2026.xlsx")

DATA_DIR = VAULT / "production" / "don-hang"
STATE_FILE = DATA_DIR / "da-xu-ly.json"
BACKUP_DIR = DATA_DIR / "sao-luu"
TUDIEN_FILE = Path(__file__).resolve().parent / "tu-dien.json"

SHEET_DON = "Đơn hàng 2026"
SHEET_SO = "TỔNG HỢP"

HANG_HEADER_DON = 4          # header ở hàng 4, dữ liệu từ hàng 5
HANG_DAU_DON = 5
HANG_DAU_SO = 5              # khối TRẢ HÀNG bắt đầu hàng 5
COT_MONTH_DEN = 637          # cột A đã có =MONTH(B) sẵn tới hàng này

DUOI_THIET_KE = {".pds", ".cnt", ".pat", ".prm", ".hcd", ".q00"}
DUOI_ANH = {".jpg", ".jpeg", ".png", ".heic", ".webp", ".bmp", ".tif", ".tiff"}
SO_BAN_LUU_GIU = 30
# 0.85: đủ để bắt 'C Thủy NA' -> thuyna, nhưng KHÔNG cho 'Mai Liên' khớp sai thành
# 'C Hà Liên' (0.80). Khách lạ thà bị chặn để anh Đức khai báo, hơn là ghi sai mã KH.
NGUONG_KHOP_KHACH = 0.85

# ============================ TIỆN ÍCH ============================


def khong_dau(s: str) -> str:
    """Bỏ dấu tiếng Việt, hạ chữ thường, gom khoảng trắng."""
    if s is None:
        return ""
    s = str(s).replace("Đ", "D").replace("đ", "d")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()


def in_dam(s: str) -> None:
    print(f"\n{s}\n" + "─" * max(12, len(s)))


def tien(x) -> str:
    try:
        return f"{float(x):,.0f}".replace(",", ".")
    except (TypeError, ValueError):
        return str(x)


def doc_json(p: Path, mac_dinh):
    if not p.exists():
        return mac_dinh
    return json.loads(p.read_text(encoding="utf-8"))


def ghi_json(p: Path, data) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def doc_ngay(v) -> dt.date | None:
    """Nhận nhiều kiểu ngày: date, datetime, 'YYYY-MM-DD', '28/07', '28/07/2026'."""
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    for f in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, f).date()
        except ValueError:
            pass
    m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})", s)      # '28/07' -> lấy năm hiện tại
    if m:
        return dt.date(dt.date.today().year, int(m.group(2)), int(m.group(1)))
    return None


def doc_so(v) -> float | None:
    """Nhận '6.500.000', '6,500,000', '2tr9', 2990000."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower().replace(" ", "")
    m = re.fullmatch(r"(\d+)tr(\d*)", s)                # '2tr9' -> 2.900.000
    if m:
        le = m.group(2)
        return float(m.group(1)) * 1_000_000 + (float(le.ljust(3, "0")) * 1000 if le else 0)
    s = re.sub(r"[^\d,.\-]", "", s)
    if "," in s and "." in s:                            # '6,500.00' kiểu Anh
        s = s.replace(",", "")
    else:
        s = s.replace(".", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def excel_dang_mo(p: Path) -> bool:
    """Excel giữ khoá bằng file tạm ~$<tên>. Có nó tức là file đang mở."""
    return (p.parent / f"~${p.name}").exists()


def luu_wb(wb, p: Path) -> None:
    """Lưu workbook, bắt Excel tính lại toàn bộ công thức khi mở.

    openpyxl giữ nguyên công thức nhưng xoá giá trị đệm; không có cờ này thì Excel
    có thể hiển thị ô rỗng cho tới khi người dùng tự bấm tính lại.
    """
    try:
        wb.calculation.fullCalcOnLoad = True
    except AttributeError:
        pass
    wb.save(p)


def sao_luu(p: Path) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    dich = BACKUP_DIR / f"{p.stem}--{dt.datetime.now():%Y%m%d-%H%M%S}{p.suffix}"
    shutil.copy2(p, dich)
    ban = sorted(BACKUP_DIR.glob(f"{p.stem}--*{p.suffix}"))
    for cu in ban[:-SO_BAN_LUU_GIU]:
        cu.unlink(missing_ok=True)
    return dich


# ============================ QUY TẮC ĐUÔI MÃ DỆT ============================
# Anh Đức xác nhận 28/07/2026:
#   "Trơn"      -> T     (Co42x7T)
#   "Kẻ"        -> K     (Co42x7K)
#   "2 kẻ chân" -> TH    (Co42x7TH)
#   còn lại     -> không đuôi, thêm " (usb)"   (Co42x7 (usb))
# BẮT BUỘC kiểm "2 kẻ chân" TRƯỚC "kẻ" vì "2 kẻ chân" cũng chứa chữ "kẻ".

def duoi_ma_det(mo_ta: str) -> tuple[str, str, str]:
    """Trả về (đuôi, hậu_tố, tên_kiểu) suy từ chữ mô tả kiểu bo."""
    s = khong_dau(mo_ta)
    if re.search(r"2\s*ke\s*chan", s):
        return "TH", "", "2 kẻ chân"
    if "tron" in s:
        return "T", "", "trơn"
    if "ke" in s:
        return "K", "", "kẻ"
    return "", " (usb)", "khác (usb)"


def dung_ma_det(ho: str, kho: str, mo_ta_kieu: str) -> tuple[str, str]:
    """Ghép mã dệt: <Họ><khổ><đuôi>[ (usb)].  Ví dụ ('Co','42x7','kẻ') -> 'Co42x7K'."""
    duoi, hau_to, ten_kieu = duoi_ma_det(mo_ta_kieu)
    ho = (ho or "").strip()
    ho = ho[:1].upper() + ho[1:].lower() if ho else ""
    kho = re.sub(r"\s+", "", (kho or "")).lower().replace("*", "x").replace("×", "x")
    return f"{ho}{kho}{duoi}{hau_to}", ten_kieu


def ma_goc(ma: str) -> str:
    """Bỏ ' (usb)' và chuẩn hoá HOA để đối chiếu với thư mục Thiết kế."""
    return re.sub(r"\s*\(usb\)\s*", "", str(ma or ""), flags=re.I).strip().upper()


# ============================ TỪ ĐIỂN ============================


def lay_tu_dien() -> dict:
    if not TUDIEN_FILE.exists():
        sys.exit("Chưa có tools/tu-dien.json. Chạy trước: python tools/dh.py xay-tu-dien")
    return doc_json(TUDIEN_FILE, {})


def lenh_xay_tu_dien(_args) -> None:
    in_dam("DỰNG TỪ ĐIỂN")

    # --- mã dệt có thật trong thư mục Thiết kế ---
    ma_det: set[str] = set()
    if THIETKE_DIR.exists():
        for f in THIETKE_DIR.rglob("*"):
            if f.is_file() and f.suffix.lower() in DUOI_THIET_KE:
                ma_det.add(f.stem.strip().upper())
    print(f"  mã dệt      : {len(ma_det):>5} mã  (từ {THIETKE_DIR.name}/)")

    # --- mã KH + mã CP từ sheet Code của sổ thật ---
    khach: dict[str, str] = {}
    dia_chi: dict[str, str] = {}
    chi_phi: dict[str, str] = {}
    if XL_SO.exists():
        ws = openpyxl.load_workbook(XL_SO, data_only=True, read_only=True)["Code"]
        rows = list(ws.iter_rows(min_row=1, max_row=110, values_only=True))
        muc = None
        for r in rows:
            cells = ["" if c is None else str(c).strip() for c in (list(r) + [""] * 6)[:6]]
            nhan = khong_dau(cells[1])
            if "ma kh" in nhan:
                muc = "kh"
                continue
            if "ma chi phi" in nhan:
                muc = "cp"
                continue
            if nhan.startswith("ma ") and "kh" not in nhan and "chi phi" not in nhan:
                muc = None
                continue
            ma, ten = cells[1].strip(), cells[2].strip()
            if not ma or not ten:
                continue
            if muc == "kh":
                khach[ma] = ten
                if cells[4].strip():
                    dia_chi[ma] = cells[4].strip()      # cột E = Địa chỉ ('Cao Bằng', 'Thanh Hóa'...)
            elif muc == "cp":
                chi_phi[ma] = ten
    print(f"  mã khách    : {len(khach):>5} mã")
    print(f"  mã chi phí  : {len(chi_phi):>5} mã")

    # --- alias khách: tên hiển thị trên Zalo -> mã KH ---
    alias: dict[str, str] = {}

    def them_alias(chuoi: str, ma: str) -> None:
        k = khong_dau(chuoi)
        if k and len(k) >= 2:
            alias.setdefault(k, ma)

    bo_tien_to = re.compile(r"^(c|co|a|anh|chi|ch|cty|cong ty|nha may|ba|chu)\s+")

    def sinh_alias(ten: str, ma: str, noi: str = "") -> None:
        """Sinh mọi biến thể anh Đức có thể viết cho một khách."""
        them_alias(ten, ma)
        goi = bo_tien_to.sub("", khong_dau(ten))
        them_alias(goi, ma)
        tu = goi.split()
        if not tu:
            return
        them_alias(tu[0], ma)                             # 'C Hoàn Cao Bằng' -> 'hoan'
        # viết tắt địa danh — anh Đức hay viết kiểu này, đúng như các mã đã có sẵn
        # trong sổ: thuyna = C Thủy Nghệ An, hangvp = C Hằng VP, phuongvp = C Phương VP
        phan_noi = khong_dau(noi).split() or tu[1:]
        if phan_noi:
            vt = "".join(t[0] for t in phan_noi)
            day_du = " ".join(phan_noi)
            for bien_the in (f"{tu[0]} {vt}", f"{tu[0]}{vt}", f"{tu[0]} {day_du}"):
                them_alias(bien_the, ma)

    for ma, ten in khach.items():
        them_alias(ma, ma)
        sinh_alias(ten, ma, dia_chi.get(ma, ""))

    # tên trong sheet THU NHẬP (cột B) khớp mã ở cột C — thường là tên đầy đủ kèm địa danh
    if XL_SO.exists():
        ws = openpyxl.load_workbook(XL_SO, data_only=True, read_only=True)["THU NHẬP"]
        for r in ws.iter_rows(min_row=3, max_row=60, min_col=2, max_col=3, values_only=True):
            ten, ma = r[0], r[1]
            if ten and ma and str(ma).strip() in khach:
                sinh_alias(str(ten), str(ma).strip())
    print(f"  alias khách : {len(alias):>5} biến thể")

    ghi_json(TUDIEN_FILE, {
        "ngay_dung": dt.date.today().isoformat(),
        "ghi_chu": "Sinh tự động bởi tools/dh.py xay-tu-dien. Anh Đức sửa tay được; "
                   "chạy lại lệnh sẽ ghi đè.",
        "quy_tac_duoi_ma": {
            "2 kẻ chân": "TH", "trơn": "T", "kẻ": "K",
            "còn lại": "không đuôi + ' (usb)'",
            "thu_tu_kiem": "PHẢI kiểm '2 kẻ chân' trước 'kẻ'",
        },
        "ma_det": sorted(ma_det),
        "khach": khach,
        "dia_chi_khach": dia_chi,
        "alias_khach": alias,
        "chi_phi": chi_phi,
    })
    print(f"\n  ✓ đã ghi {TUDIEN_FILE.relative_to(VAULT)}")

    for thu in ("GA90X12", "CO40X7K", "GDO80X16", "GA50X12", "GA80X9"):
        print(f"    {thu:<10} {'CÓ' if thu in ma_det else 'KHÔNG CÓ'}")


def tim_ma_kh(ten: str, td: dict) -> tuple[str | None, float, str]:
    """Khớp tên khách -> mã KH. Trả về (mã, độ_khớp, cách_khớp)."""
    if not ten:
        return None, 0.0, "rỗng"
    k = khong_dau(ten)
    alias, khach = td.get("alias_khach", {}), td.get("khach", {})
    if k in khach:
        return k, 1.0, "đúng mã"
    if k in alias:
        return alias[k], 1.0, "đúng alias"
    bo = re.sub(r"^(c|co|a|anh|chi|ch|cty|cong ty|nha may|ba|chu)\s+", "", k)
    if bo in alias:
        return alias[bo], 0.97, "alias sau khi bỏ tiền tố"
    ung_vien = list(alias.keys())
    for ngu in (k, bo):
        got = difflib.get_close_matches(ngu, ung_vien, n=1, cutoff=NGUONG_KHOP_KHACH)
        if got:
            return alias[got[0]], difflib.SequenceMatcher(None, ngu, got[0]).ratio(), f"khớp mờ '{got[0]}'"
    return None, 0.0, "không tìm được"


def lenh_them_alias(args) -> None:
    """Khai báo tên khách mới cho từ điển, không cần dựng lại toàn bộ."""
    in_dam("THÊM ALIAS KHÁCH")
    td = lay_tu_dien()
    ma = args.ma_kh.strip()
    if ma not in td.get("khach", {}):
        print(f"  ⚠ mã '{ma}' CHƯA có trong sheet Code của sổ thật.")
        print("    Anh Đức cần thêm mã này vào sheet 'Code' (khối Mã KH) và sheet 'THU NHẬP',")
        print("    nếu không thì doanh thu của khách này sẽ không được cộng vào bảng.")
        print(f"    Các mã đang có gần giống: "
              f"{difflib.get_close_matches(ma, list(td.get('khach', {})), n=5, cutoff=0.4)}")
        if not args.ep:
            sys.exit("  Dừng. Thêm --ep nếu vẫn muốn ghi alias.")
    k = khong_dau(args.ten)
    cu = td.setdefault("alias_khach", {}).get(k)
    if cu and cu != ma:
        print(f"  ⚠ alias '{k}' đang trỏ tới '{cu}', sẽ đổi thành '{ma}'")
    td["alias_khach"][k] = ma
    ghi_json(TUDIEN_FILE, td)
    print(f"  ✓ '{args.ten}'  →  mã '{ma}'  ({td.get('khach', {}).get(ma, 'chưa có tên')})")
    print("  Lưu ý: chạy 'xay-tu-dien' sẽ ghi đè — khai báo lại sau khi chạy lệnh đó.")


# ============================ ẢNH ============================


def bam_file(p: Path) -> str:
    h = hashlib.sha1()
    with p.open("rb") as f:
        for khoi in iter(lambda: f.read(1 << 20), b""):
            h.update(khoi)
    return h.hexdigest()


def da_doc_roi(h: str, da_co) -> bool:
    """Ảnh đã đọc chưa?

    'anh-moi' in ra hash RÚT GỌN 10 ký tự và LLM chép đúng chuỗi đó vào JSON, nên khoá
    trong da-xu-ly.json là tiền tố chứ không phải hash đầy đủ. So sánh thẳng h với khoá
    sẽ không bao giờ khớp — phải so theo tiền tố.
    """
    return any(h.startswith(k) for k in da_co)


def lenh_anh_moi(args) -> None:
    in_dam("ẢNH CHƯA XỬ LÝ")
    if not ANH_DIR.exists():
        sys.exit(f"Không thấy thư mục ảnh: {ANH_DIR}")

    state = doc_json(STATE_FILE, {})
    da_co = set(state.get("anh_da_doc", {}))
    anh = sorted(p for p in ANH_DIR.rglob("*") if p.is_file() and p.suffix.lower() in DUOI_ANH)
    moi = [(p, h) for p, h in ((p, bam_file(p)) for p in anh) if not da_doc_roi(h, da_co)]

    print(f"  tổng ảnh trong thư mục : {len(anh)}")
    print(f"  đã đọc lần trước       : {len(da_co)}")
    print(f"  CẦN ĐỌC                : {len(moi)}")
    if not moi:
        print("\n  Không có ảnh mới. Dừng — không tốn token.")
        return

    ra_dir = Path(args.ra) if args.ra else (DATA_DIR / "anh-thu-nho")
    ra_dir.mkdir(parents=True, exist_ok=True)
    try:
        from PIL import Image, ImageOps
    except ImportError:
        print("\n  (Thiếu Pillow — bỏ qua thu nhỏ, dùng ảnh gốc)")
        Image = None

    print()
    for p, h in moi:
        dich = p
        if Image is not None:
            try:
                with Image.open(p) as im:
                    im = ImageOps.exif_transpose(im).convert("L")
                    im.thumbnail((args.canh, args.canh), Image.LANCZOS)
                    im = ImageOps.autocontrast(im, cutoff=1)
                    dich = ra_dir / f"{h[:10]}--{p.stem}.jpg"
                    im.save(dich, "JPEG", quality=82, optimize=True)
            except Exception as e:                                  # noqa: BLE001
                print(f"  ! không thu nhỏ được {p.name}: {e}")
                dich = p
        print(f"  {dich}")
        print(f"      (gốc: {p.name}  sha1={h[:10]})")

    print("\n  Đọc các đường dẫn trên, rồi ghi JSON và chạy: dh.py don-them --json <file>")
    print(f"  Nhớ điền \"anh_sha1\" cho mỗi đơn: {', '.join(h[:10] for _, h in moi)}")


# ============================ BƯỚC 3: GHI ĐƠN HÀNG ============================


def _kieu_dinh_dang(ws, hang_mau: int, cot: int):
    o = ws.cell(row=hang_mau, column=cot)
    return o.number_format, o.font.name, o.font.size


def lenh_don_them(args) -> None:
    in_dam("BƯỚC 3 — GHI ĐƠN HÀNG VÀO EXCEL")
    td = lay_tu_dien()
    goi = doc_json(Path(args.json), None)
    if goi is None:
        sys.exit(f"Không đọc được {args.json}")
    if not XL_DON.exists():
        sys.exit(f"Không thấy {XL_DON}")
    if excel_dang_mo(XL_DON) and not args.thu:
        sys.exit(f"File đang mở trong Excel: {XL_DON.name}\n  → Anh Đức đóng file rồi chạy lại.")

    wb = openpyxl.load_workbook(XL_DON)
    if SHEET_DON not in wb.sheetnames:
        sys.exit(f"Không thấy sheet '{SHEET_DON}'. Chạy: python tools/dh.py sua-file-don")
    ws = wb[SHEET_DON]

    # --- đọc các dòng đã có: để chống trùng và cấp mã đơn tiếp theo ---
    da_co_khoa: set[tuple] = set()
    ma_don_da_co: set[str] = set()
    so_ms_lon_nhat = 0
    hang_ke = HANG_DAU_DON
    for r in range(HANG_DAU_DON, ws.max_row + 1):
        ngay, khach, md = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value
        if all(ws.cell(r, c).value in (None, "") for c in range(1, 13)):
            continue
        hang_ke = r + 1
        d = doc_ngay(ngay)
        da_co_khoa.add((
            d.isoformat() if d else "",
            khong_dau(ws.cell(r, 12).value or khach),
            ma_goc(ws.cell(r, 6).value),
            khong_dau(ws.cell(r, 7).value),
            doc_so(ws.cell(r, 9).value),
        ))
        if md:
            ma_don_da_co.add(str(md).strip())
            m = re.fullmatch(r"[Mm][Ss](\d+)", str(md).strip())
            if m:
                so_ms_lon_nhat = max(so_ms_lon_nhat, int(m.group(1)))

    # mã Ms đã dùng trong sổ cũ (bang-so-van-hanh) tối thiểu tới Ms60
    so_ms_lon_nhat = max(so_ms_lon_nhat, int(goi.get("ms_bat_dau_tu", 60)))
    stt = max(
        (int(ws.cell(r, 1).value) for r in range(HANG_DAU_DON, hang_ke)
         if isinstance(ws.cell(r, 1).value, (int, float))),
        default=0,
    )

    dinh_dang = {c: _kieu_dinh_dang(ws, HANG_HEADER_DON, c) for c in range(1, 13)}

    se_ghi: list[list] = []
    canh_bao: list[str] = []
    bo_trung: list[str] = []
    anh_dung: dict[str, str] = {}

    for don in goi.get("don", []):
        ngay = doc_ngay(don.get("ngay"))
        if ngay is None:
            canh_bao.append(f"BỎ ĐƠN — không đọc được ngày: {don.get('ngay')!r} (khách {don.get('khach')})")
            continue
        ten_khach = str(don.get("khach") or "").strip()
        ma_kh = (don.get("ma_kh") or "").strip()
        if ma_kh and ma_kh in td.get("khach", {}):
            do_khop, cach = 1.0, "do LLM cấp"
        else:
            ma_kh, do_khop, cach = tim_ma_kh(ten_khach, td)
        if not ma_kh:
            canh_bao.append(f"BỎ ĐƠN — không nhận ra khách {ten_khach!r} ({cach}). "
                            f"Thêm alias vào tools/tu-dien.json rồi chạy lại.")
            continue
        if do_khop < 0.95:
            canh_bao.append(f"khách {ten_khach!r} → mã '{ma_kh}' ({cach}, khớp {do_khop:.0%}) — anh Đức soát lại")

        ma_don = str(don.get("ma_don") or "").strip()
        if ma_don and ma_don in ma_don_da_co:
            bo_trung.append(f"{ma_don} (mã đơn đã có trong sheet)")
            continue
        nhan = ma_don or f"đơn {ten_khach} {ngay:%d/%m}"

        hang_don = don.get("hang") or []
        if not hang_don:
            canh_bao.append(f"BỎ ĐƠN {nhan} — không có dòng hàng nào")
            continue

        # Gom các dòng hợp lệ TRƯỚC, chỉ cấp mã Ms sau khi biết đơn có dòng sống sót —
        # nếu không thì đơn trùng hoàn toàn vẫn ngốn mất một số Ms.
        hang_ok: list[list] = []
        for h in hang_don:
            ma_det = str(h.get("ma_det") or "").strip()
            ten_kieu = h.get("kieu") or ""
            thieu_kho = False
            if not ma_det:
                if not str(h.get("kho") or "").strip():
                    # Anh Đức chọn: thiếu khổ thì vẫn ghi đơn để không sót, ô Mẫu dệt để trống
                    # và đánh dấu THIẾU KHỔ; 'soat' sẽ nhắc đến khi anh điền xong.
                    thieu_kho, ma_det, ten_kieu = True, "", (h.get("kieu") or "")
                else:
                    ma_det, ten_kieu = dung_ma_det(h.get("ho"), h.get("kho"), h.get("kieu") or h.get("mo_ta"))
            if not thieu_kho and not ma_goc(ma_det):
                canh_bao.append(f"BỎ DÒNG {nhan} — không dựng được mã dệt từ {h!r}")
                continue

            goc = ma_goc(ma_det)
            la_usb = bool(re.search(r"\(usb\)", ma_det, re.I))
            if thieu_kho:
                canh_bao.append(f"{nhan} · '{h.get('mo_ta') or ''}' — THIẾU KHỔ, ô Mẫu dệt để trống. "
                                f"Anh Đức điền khổ rồi em dựng mã sau.")
            elif goc not in set(td.get("ma_det", [])):
                if la_usb:
                    canh_bao.append(f"{nhan} · {ma_det} — không có file thiết kế (đúng kỳ vọng cho hàng USB)")
                else:
                    canh_bao.append(f"BỎ DÒNG {nhan} · {ma_det} — KHÔNG có file thiết kế trong Thiết kế/. "
                                    f"Kiểm tra lại mã.")
                    continue

            sl = doc_so(h.get("so_luong"))
            if sl is None or sl <= 0:
                canh_bao.append(f"BỎ DÒNG {nhan} · {ma_det} — số lượng không hợp lệ: {h.get('so_luong')!r}")
                continue

            mau = str(h.get("mau") or "").strip()
            khoa = (ngay.isoformat(), khong_dau(ma_kh), goc, khong_dau(mau), sl)
            if khoa in da_co_khoa:
                bo_trung.append(f"{nhan} · {ma_det} · {mau} · {sl:g} (đã có trong sheet)")
                continue
            da_co_khoa.add(khoa)

            gc = str(don.get("ghi_chu") or h.get("ghi_chu") or "").strip()
            if thieu_kho:
                gc = ("⚠ THIẾU KHỔ — cần điền" + (f" · {gc}" if gc else ""))
            hang_ok.append([
                None, ngay, ten_khach or td["khach"].get(ma_kh, ""), None,
                str(h.get("mo_ta") or "").strip(), ma_det, mau,
                str(h.get("kich_thuoc") or "").strip(), sl,
                str(h.get("dvt") or "").strip().lower(), gc, ma_kh,
            ])

        if not hang_ok:
            continue
        if not ma_don:
            so_ms_lon_nhat += 1
            ma_don = f"Ms{so_ms_lon_nhat}"
        ma_don_da_co.add(ma_don)
        for row in hang_ok:
            stt += 1
            row[0], row[3] = stt, ma_don
            se_ghi.append(row)
        if don.get("anh_sha1"):
            anh_dung[str(don["anh_sha1"])] = ma_don

    # ------------------------- báo cáo -------------------------
    print(f"  ghi vào : {XL_DON.name} · sheet '{SHEET_DON}' · từ hàng {hang_ke}")
    print(f"  dòng mới: {len(se_ghi)}   bỏ vì trùng: {len(bo_trung)}   cảnh báo: {len(canh_bao)}")

    if se_ghi:
        print("\n  ┌─ SẼ GHI " + "─" * 84)
        print(f"  │ {'Mã đơn':<7}{'Ngày':<11}{'Mã KH':<12}{'Mã dệt':<16}{'Màu':<26}{'SL':>7} {'ĐVT'}")
        for r in se_ghi:
            print(f"  │ {r[3]:<7}{r[1]:%d/%m/%Y} {r[11]:<12}{r[5]:<16}{r[6][:25]:<26}{r[8]:>7g} {r[9]}")
        print("  └" + "─" * 92)
    if bo_trung:
        print("\n  BỎ VÌ TRÙNG:")
        for x in bo_trung:
            print(f"    · {x}")
    if canh_bao:
        print("\n  ⚠ CẢNH BÁO:")
        for x in canh_bao:
            print(f"    · {x}")

    if args.thu:
        print("\n  [CHẠY KHÔ] Không ghi gì. Bỏ --thu để ghi thật.")
        return
    if not se_ghi:
        print("\n  Không có gì để ghi.")
        return

    print(f"\n  sao lưu → {sao_luu(XL_DON).name}")
    for i, row in enumerate(se_ghi):
        r = hang_ke + i
        for c, val in enumerate(row, start=1):
            o = ws.cell(row=r, column=c, value=val)
            _nf, ten_font, co_font = dinh_dang[c]
            o.font = Font(name=ten_font, size=co_font)
            o.alignment = Alignment(vertical="center", wrap_text=(c in (5, 7, 11)))
            if c == 2:
                o.number_format = "DD/MM/YYYY"
            elif c == 9:
                o.number_format = "#,##0"
    luu_wb(wb, XL_DON)
    print(f"  ✓ đã ghi {len(se_ghi)} dòng vào {XL_DON.name}")

    state = doc_json(STATE_FILE, {})
    state.setdefault("anh_da_doc", {})
    for h, md in anh_dung.items():
        state["anh_da_doc"][h] = {"ma_don": md, "ngay_doc": dt.date.today().isoformat()}
    state["lan_cuoi_don"] = dt.datetime.now().isoformat(timespec="seconds")
    ghi_json(STATE_FILE, state)
    print(f"  ✓ đã ghi dấu {len(anh_dung)} ảnh vào {STATE_FILE.name} — sẽ không đọc lại")


# ============================ BƯỚC 6: GHI TRẢ HÀNG ============================


def lenh_tra_them(args) -> None:
    in_dam("BƯỚC 6 — GHI TRẢ HÀNG VÀO SỔ THẬT")
    td = lay_tu_dien()
    goi = doc_json(Path(args.json), None)
    if goi is None:
        sys.exit(f"Không đọc được {args.json}")
    if not XL_SO.exists():
        sys.exit(f"Không thấy {XL_SO}")
    if excel_dang_mo(XL_SO) and not args.thu:
        sys.exit(f"File đang mở trong Excel: {XL_SO.name}\n  → Anh Đức đóng file rồi chạy lại.")

    wb = openpyxl.load_workbook(XL_SO)
    ws = wb[SHEET_SO]

    da_co: set[tuple] = set()
    hang_ke = HANG_DAU_SO
    for r in range(HANG_DAU_SO, ws.max_row + 1):
        ngay, ma_kh, nd, st = (ws.cell(r, c).value for c in (2, 3, 4, 5))
        if ngay in (None, "") and ma_kh in (None, ""):
            continue
        hang_ke = r + 1
        d = doc_ngay(ngay)
        # KHÔNG đưa nội dung vào khoá: cùng một lần trả hàng, anh Đức viết trong sổ và
        # viết trên Zalo khác nhau ('xanh biển 150b' vs 'xanh biển 150 bộ') — đưa nội dung
        # vào khoá thì không bắt được trùng và tiền bị cộng đôi.
        da_co.add((d.isoformat() if d else "", khong_dau(ma_kh), doc_so(st)))

    mau_ngay = ws.cell(hang_ke - 1, 2)
    mau_tien = ws.cell(hang_ke - 1, 5)
    nf_ngay = mau_ngay.number_format
    nf_tien = mau_tien.number_format

    se_ghi: list[list] = []
    canh_bao: list[str] = []
    bo_trung: list[str] = []

    for t in goi.get("tra", []):
        ngay = doc_ngay(t.get("ngay"))
        if ngay is None:
            canh_bao.append(f"BỎ — không đọc được ngày {t.get('ngay')!r} ({t.get('khach')})")
            continue
        if ngay > dt.date.today():
            canh_bao.append(f"BỎ — ngày ở tương lai: {ngay:%d/%m/%Y} ({t.get('khach')})")
            continue

        ten = str(t.get("khach") or "").strip()
        ma_kh = (t.get("ma_kh") or "").strip()
        if ma_kh and ma_kh in td.get("khach", {}):
            do_khop, cach = 1.0, "do LLM cấp"
        else:
            ma_kh, do_khop, cach = tim_ma_kh(ten, td)
        if not ma_kh:
            canh_bao.append(f"BỎ — không nhận ra khách {ten!r} ({cach}). Thêm alias vào tu-dien.json.")
            continue
        if do_khop < 0.95:
            canh_bao.append(f"khách {ten!r} → '{ma_kh}' ({cach}, khớp {do_khop:.0%}) — anh Đức soát lại")

        nd = str(t.get("noi_dung") or "").strip()
        if not nd:
            canh_bao.append(f"BỎ — thiếu nội dung ({ten}, {ngay:%d/%m})")
            continue

        st = doc_so(t.get("thanh_tien"))
        if st is None or st <= 0:
            canh_bao.append(f"BỎ — thành tiền không hợp lệ {t.get('thanh_tien')!r} ({ten}, {nd})")
            continue

        khoa = (ngay.isoformat(), khong_dau(ma_kh), st)
        if khoa in da_co:
            bo_trung.append(f"{ngay:%d/%m} · {ma_kh} · {tien(st)} đ — sổ đã có (nội dung có thể "
                            f"viết khác: '{nd}')")
            continue
        da_co.add(khoa)
        se_ghi.append([ngay, ma_kh, nd, st])

    se_ghi.sort(key=lambda x: x[0])
    tong = sum(r[3] for r in se_ghi)

    print(f"  ghi vào : {XL_SO.name} · sheet '{SHEET_SO}' · khối TRẢ HÀNG cột B–E · từ hàng {hang_ke}")
    print(f"  dòng mới: {len(se_ghi)}   bỏ vì trùng: {len(bo_trung)}   cảnh báo: {len(canh_bao)}")
    if se_ghi:
        print("\n  ┌─ SẼ GHI " + "─" * 76)
        print(f"  │ {'Ngày':<12}{'Mã KH':<14}{'Nội dung':<38}{'Thành tiền':>13}")
        for r in se_ghi:
            print(f"  │ {r[0]:%d/%m/%Y}  {r[1]:<14}{r[2][:37]:<38}{tien(r[3]):>13}")
        print(f"  │ {'':<12}{'':<14}{'TỔNG':<38}{tien(tong):>13}")
        print("  └" + "─" * 84)
    if bo_trung:
        print("\n  BỎ VÌ TRÙNG:")
        for x in bo_trung:
            print(f"    · {x}")
    if canh_bao:
        print("\n  ⚠ CẢNH BÁO:")
        for x in canh_bao:
            print(f"    · {x}")

    if args.thu:
        print("\n  [CHẠY KHÔ] Không ghi gì. Bỏ --thu để ghi thật.")
        return
    if not se_ghi:
        print("\n  Không có gì để ghi.")
        return

    print(f"\n  sao lưu → {sao_luu(XL_SO).name}")
    for i, (ngay, ma_kh, nd, st) in enumerate(se_ghi):
        r = hang_ke + i
        o = ws.cell(r, 2, dt.datetime(ngay.year, ngay.month, ngay.day)); o.number_format = nf_ngay
        ws.cell(r, 3, ma_kh)
        ws.cell(r, 4, nd)
        o = ws.cell(r, 5, st); o.number_format = nf_tien
        # cột A có sẵn công thức tháng; nếu ghi quá vùng đó thì bù thêm.
        # Dùng dạng IF để ô ngày rỗng không sinh ra số 1 rác (xem chuan-hoa-cong-thuc).
        if r > COT_MONTH_DEN and ws.cell(r, 1).value in (None, ""):
            ws.cell(r, 1, f'=IF(B{r}="","",MONTH(B{r}))')
    luu_wb(wb, XL_SO)
    print(f"  ✓ đã ghi {len(se_ghi)} dòng, tổng {tien(tong)} đ")

    state = doc_json(STATE_FILE, {})
    moc = max(r[0] for r in se_ghi).isoformat()
    if moc > state.get("moc_tra_hang", ""):
        state["moc_tra_hang"] = moc
    state["lan_cuoi_tra"] = dt.datetime.now().isoformat(timespec="seconds")
    ghi_json(STATE_FILE, state)
    print(f"  ✓ mốc đã đọc nhóm Zalo 'Trả hàng' → {moc}")
    print("\n  Anh Đức mở file kiểm sheet LỢI NHUẬN tháng "
          f"{se_ghi[-1][0].month} xem có tăng đúng {tien(tong)} đ không.")


def lenh_sua_kho(args) -> None:
    """Điền khổ cho các dòng bị đánh dấu THIẾU KHỔ của một đơn, rồi dựng lại mã dệt."""
    in_dam("ĐIỀN KHỔ CHO ĐƠN THIẾU")
    td = lay_tu_dien()
    if excel_dang_mo(XL_DON) and not args.thu:
        sys.exit(f"File đang mở trong Excel: {XL_DON.name} → đóng rồi chạy lại.")

    wb = openpyxl.load_workbook(XL_DON)
    ws = wb[SHEET_DON]
    ma_don = args.ma_don.strip()

    se_sua: list[tuple[int, str, str]] = []
    for r in range(HANG_DAU_DON, ws.max_row + 1):
        if str(ws.cell(r, 4).value or "").strip() != ma_don:
            continue
        if ws.cell(r, 6).value:                       # đã có mã dệt rồi thì bỏ qua
            continue
        # Kiểu suy từ ô 'Tên mẫu / Mô tả' — đúng chữ anh Đức viết trong cú pháp.
        mo_ta = str(ws.cell(r, 5).value or "")
        ma_det, ten_kieu = dung_ma_det(args.ho, args.kho, mo_ta)
        goc = ma_goc(ma_det)
        if goc not in set(td.get("ma_det", [])) and "(usb)" not in ma_det.lower():
            sys.exit(f"Mã dựng ra là '{ma_det}' nhưng KHÔNG có file thiết kế trong Thiết kế/.\n"
                     f"  Kiểm lại họ mã / khổ. Các mã gần giống: "
                     f"{difflib.get_close_matches(goc, td.get('ma_det', []), n=6, cutoff=0.6)}")
        se_sua.append((r, ma_det, ten_kieu))

    if not se_sua:
        print(f"  Đơn {ma_don} không có dòng nào đang trống ô Mẫu dệt.")
        return

    print(f"  đơn {ma_don} · họ '{args.ho}' · khổ '{args.kho}'")
    for r, ma_det, ten_kieu in se_sua:
        print(f"    hàng {r}: '{ws.cell(r, 5).value}'  →  {ma_det}   (kiểu: {ten_kieu})")
    if args.thu:
        print("\n  [CHẠY KHÔ] Không ghi gì. Bỏ --thu để ghi thật.")
        return

    print(f"\n  sao lưu → {sao_luu(XL_DON).name}")
    for r, ma_det, _ in se_sua:
        ws.cell(r, 6, ma_det)
        if not ws.cell(r, 8).value:
            ws.cell(r, 8, args.kho)
        gc = str(ws.cell(r, 11).value or "")
        gc = re.sub(r"⚠ THIẾU KHỔ — cần điền(\s*·\s*)?", "", gc).strip(" ·")
        ws.cell(r, 11, gc or None)
    luu_wb(wb, XL_DON)
    print(f"  ✓ đã điền {len(se_sua)} dòng và gỡ dấu THIẾU KHỔ")


# ============================ RÀ SOÁT ============================


def lenh_xoa_tra(args) -> None:
    """Xoá các dòng trong khối TRẢ HÀNG theo số hàng (dùng khi lỡ ghi trùng)."""
    in_dam("XOÁ DÒNG TRẢ HÀNG")
    if excel_dang_mo(XL_SO) and not args.thu:
        sys.exit(f"File đang mở trong Excel: {XL_SO.name} → đóng rồi chạy lại.")
    wb = openpyxl.load_workbook(XL_SO)
    ws = wb[SHEET_SO]

    hang = sorted({int(h) for h in args.hang}, reverse=True)
    se_xoa = []
    for r in hang:
        if r < HANG_DAU_SO:
            sys.exit(f"Hàng {r} nằm trên vùng dữ liệu (bắt đầu từ {HANG_DAU_SO}).")
        d, ma_kh, nd, st = (ws.cell(r, c).value for c in (2, 3, 4, 5))
        if d in (None, "") and ma_kh in (None, ""):
            print(f"  hàng {r} đã trống — bỏ qua")
            continue
        se_xoa.append((r, doc_ngay(d), ma_kh, nd, doc_so(st)))

    if not se_xoa:
        print("  Không có gì để xoá.")
        return
    print(f"  {'Hàng':<7}{'Ngày':<12}{'Mã KH':<14}{'Nội dung':<46}{'Thành tiền':>13}")
    for r, d, ma_kh, nd, st in se_xoa:
        print(f"  {r:<7}{d:%d/%m/%Y}  {str(ma_kh):<14}{str(nd)[:45]:<46}{tien(st):>13}")
    print(f"  → tổng sẽ giảm {tien(sum(x[4] or 0 for x in se_xoa))} đ")

    if args.thu:
        print("\n  [CHẠY KHÔ] Không xoá gì. Bỏ --thu để xoá thật.")
        return

    print(f"\n  sao lưu → {sao_luu(XL_SO).name}")
    # Chỉ XOÁ NỘI DUNG 4 ô B–E, không delete_rows: cột A có sẵn công thức =MONTH(B)
    # trải tới hàng 637, xoá cả hàng sẽ làm lệch toàn bộ vùng đó.
    for r, *_ in se_xoa:
        for c in (2, 3, 4, 5):
            ws.cell(r, c).value = None
    luu_wb(wb, XL_SO)
    print(f"  ✓ đã xoá {len(se_xoa)} dòng")
    print("  Lưu ý: các ô để trống tại chỗ, dòng ghi sau sẽ nối tiếp bên dưới.")


def _chep_dinh_dang(nguon, dich) -> None:
    """Sao chép nguyên định dạng ô (font, màu, viền, number_format, căn lề).

    Phải chép cả StyleArray chứ không đặt lại từng thuộc tính: màu chữ trong sổ của anh
    Đức dùng theme/indexed color, dựng Font() mới sẽ làm mất màu gốc.
    """
    dich._style = _copy.copy(nguon._style)


def lenh_don_tra(args) -> None:
    """Dồn khối TRẢ HÀNG lên, lấp các hàng trống ở giữa (sau khi xoá dòng trùng)."""
    in_dam("DỒN KHỐI TRẢ HÀNG")
    if excel_dang_mo(XL_SO) and not args.thu:
        sys.exit(f"File đang mở trong Excel: {XL_SO.name} → đóng rồi chạy lại.")
    wb = openpyxl.load_workbook(XL_SO)
    ws = wb[SHEET_SO]

    cuoi = max((r for r in range(HANG_DAU_SO, ws.max_row + 1)
                if ws.cell(r, 2).value not in (None, "")), default=0)
    con = [r for r in range(HANG_DAU_SO, cuoi + 1) if ws.cell(r, 2).value not in (None, "")]
    trong = [r for r in range(HANG_DAU_SO, cuoi + 1) if ws.cell(r, 2).value in (None, "")]
    if not trong:
        print("  Khối TRẢ HÀNG không có hàng trống ở giữa — không cần dồn.")
        return
    print(f"  hàng trống ở giữa: {trong}")
    print(f"  {len(con)} dòng có dữ liệu sẽ dồn về {HANG_DAU_SO}..{HANG_DAU_SO + len(con) - 1}")
    if args.thu:
        print("\n  [CHẠY KHÔ] Không ghi gì.")
        return

    print(f"  sao lưu → {sao_luu(XL_SO).name}")
    # Chỉ đụng cột B–E của khối TRẢ HÀNG. KHÔNG dùng delete_rows: cột A có công thức
    # =MONTH(B) trải tới hàng 637, và khối CHI PHÍ (L–P) nằm cùng hàng nhưng độc lập —
    # xoá cả hàng sẽ kéo lệch cả hai thứ đó.
    giu = [[ws.cell(r, c).value for c in (2, 3, 4, 5)] for r in con]
    kieu = [[ws.cell(r, c) for c in (2, 3, 4, 5)] for r in con]
    for i, (gt, mau) in enumerate(zip(giu, kieu)):
        r = HANG_DAU_SO + i
        for j, c in enumerate((2, 3, 4, 5)):
            o = ws.cell(r, c)
            _chep_dinh_dang(mau[j], o)
            o.value = gt[j]
    for r in range(HANG_DAU_SO + len(con), cuoi + 1):
        for c in (2, 3, 4, 5):
            ws.cell(r, c).value = None
    luu_wb(wb, XL_SO)
    print(f"  ✓ đã dồn, dòng cuối nay là {HANG_DAU_SO + len(con) - 1} (trước là {cuoi})")


def lenh_dong_bo_thu_nhap(args) -> None:
    """Đánh lại STT và đồng bộ định dạng toàn bộ vùng khách của sheet THU NHẬP."""
    in_dam("ĐỒNG BỘ SHEET THU NHẬP")
    if excel_dang_mo(XL_SO) and not args.thu:
        sys.exit(f"File đang mở trong Excel: {XL_SO.name} → đóng rồi chạy lại.")
    wb = openpyxl.load_workbook(XL_SO)
    tn = wb["THU NHẬP"]

    hang_tong = _tim_hang_tong(tn)
    if hang_tong is None:
        sys.exit("Không tìm thấy khối tổng trong THU NHẬP.")
    khach = [r for r in range(THU_NHAP_HANG_DAU, hang_tong) if tn.cell(r, 3).value]
    if not khach:
        sys.exit("Không thấy dòng khách nào.")

    # Hàng mẫu = dòng khách "sạch" nhất: dòng áp chót, chưa từng bị ghi đè lên hàng tổng cũ.
    mau = int(args.mau) if args.mau else khach[-3] if len(khach) >= 3 else khach[0]
    print(f"  {len(khach)} dòng khách (r{khach[0]}–r{khach[-1]}) · khối tổng r{hang_tong}")
    print(f"  lấy hàng {mau} làm mẫu định dạng")

    lech_stt = [(r, tn.cell(r, 1).value) for i, r in enumerate(khach, start=1)
                if tn.cell(r, 1).value != i]
    print(f"  STT sai/lệch kiểu: {len(lech_stt)} ô" + (f" → {lech_stt[:6]}" if lech_stt else ""))
    if args.thu:
        print("\n  [CHẠY KHÔ] Không ghi gì.")
        return

    print(f"  sao lưu → {sao_luu(XL_SO).name}")
    cot = list(range(1, COT_TONG_NAM + 1))
    for i, r in enumerate(khach, start=1):
        for c in cot:
            if r != mau:
                _chep_dinh_dang(tn.cell(mau, c), tn.cell(r, c))
        tn.cell(r, 1).value = i                       # STT là SỐ, không phải chuỗi
    # Hàng tổng: cùng phông chữ với vùng khách nhưng in đậm
    for c in cot:
        o = tn.cell(hang_tong, c)
        _chep_dinh_dang(tn.cell(mau, c), o)
        ft = o.font
        o.font = Font(name=ft.name, size=ft.sz, bold=True, color=ft.color)
    for c in (6, 9, 12, 15):                          # hàng cộng theo quý
        o = tn.cell(hang_tong + 1, c)
        _chep_dinh_dang(tn.cell(mau, c), o)
        ft = o.font
        o.font = Font(name=ft.name, size=ft.sz, bold=True, color=ft.color)
    luu_wb(wb, XL_SO)
    print(f"  ✓ đánh lại STT 1–{len(khach)} và đồng bộ định dạng {len(khach)} dòng khách "
          f"+ hàng TỔNG r{hang_tong} + hàng quý r{hang_tong + 1}")


def lenh_chuan_hoa_cong_thuc(args) -> None:
    """Sửa hai lỗi công thức gốc trong sổ thật.

    1. Cột A (tháng của TRẢ HÀNG) và cột L (tháng của CHI PHÍ) dùng `=MONTH(B)`. Ở các
       hàng chưa có dữ liệu, ô ngày rỗng nên `MONTH(0)` trả về **1** — sinh ra hàng trăm
       số 1 rác. Đổi thành `=IF(B="","",MONTH(B))`.
    2. Sheet THU NHẬP có 116 ô dùng `SUMIFS(..., $A:$A, lookup(D$2, $A:$A))`. `LOOKUP`
       đòi vector phải sắp xếp tăng dần; đám số 1 rác ở trên phá vỡ điều đó nên kết quả
       không xác định. Đổi sang dạng phẳng `SUMIFS(..., $A:$A, D$2)` — ngắn và đúng.
    """
    in_dam("CHUẨN HOÁ CÔNG THỨC SỔ THẬT")
    if excel_dang_mo(XL_SO) and not args.thu:
        sys.exit(f"File đang mở trong Excel: {XL_SO.name} → đóng rồi chạy lại.")
    wb = openpyxl.load_workbook(XL_SO)
    ws, tn = wb[SHEET_SO], wb["THU NHẬP"]

    sua_month = []
    for cot_thang, cot_ngay in ((1, 2), (12, 13)):
        col_ngay = ws.cell(1, cot_ngay).column_letter
        for r in range(HANG_DAU_SO, ws.max_row + 1):
            o = ws.cell(r, cot_thang)
            if isinstance(o.value, str) and o.value.upper().startswith("=MONTH("):
                sua_month.append((r, cot_thang, f'=IF({col_ngay}{r}="","",MONTH({col_ngay}{r}))'))

    sua_lookup = []
    for r in range(THU_NHAP_HANG_DAU, 200):
        for c in THANG_COT:
            v = tn.cell(r, c).value
            if isinstance(v, str) and "lookup" in v.lower() and "SUMIFS" in v.upper():
                col = tn.cell(r, c).column_letter
                sua_lookup.append((r, c, f"=SUMIFS('{SHEET_SO}'!$E:$E,'{SHEET_SO}'!$C:$C,"
                                         f"$C{r},'{SHEET_SO}'!$A:$A,{col}$2)"))

    print(f"  cột A/L — ô =MONTH(...) sẽ đổi sang =IF(...): {len(sua_month)}")
    print(f"  THU NHẬP — ô lookup() sẽ đổi sang dạng phẳng : {len(sua_lookup)}")
    if not sua_month and not sua_lookup:
        print("  Đã chuẩn hoá trước đó — không làm gì.")
        return
    if args.thu:
        print("\n  [CHẠY KHÔ] Không ghi gì.")
        return

    print(f"  sao lưu → {sao_luu(XL_SO).name}")
    for r, c, ct in sua_month:
        ws.cell(r, c).value = ct
    for r, c, ct in sua_lookup:
        tn.cell(r, c).value = ct
    luu_wb(wb, XL_SO)
    print(f"  ✓ đã sửa {len(sua_month)} ô cột tháng và {len(sua_lookup)} ô SUMIFS")
    print("  Anh Đức mở file kiểm dòng 'c Giang' — tháng 4 phải ra 30.000.000 đ.")


def _dat_hang_tong(tn, ln, cp, hang_cu: int, hang_moi: int) -> int:
    """Dời khối tổng của THU NHẬP tới hàng mới và trỏ lại 13 công thức bên ngoài."""
    if hang_moi == hang_cu:
        return 0
    for c in THANG_COT + [COT_TONG_NAM]:
        col = tn.cell(hang_cu, c).column_letter
        cu = tn.cell(hang_cu, c)
        moi = tn.cell(hang_moi, c, f"=SUM({col}{THU_NHAP_HANG_DAU}:{col}{hang_moi - 1})")
        moi.number_format, moi.font = cu.number_format, Font(bold=True)
        cu.value = None
    tn.cell(hang_moi, 2, "TỔNG").font = Font(bold=True)
    tn.cell(hang_cu, 2).value = None

    for cot_dich, (a, b) in ((6, (4, 6)), (9, (7, 9)), (12, (10, 12)), (15, (13, 15))):
        ca = tn.cell(hang_moi, a).column_letter
        cb = tn.cell(hang_moi, b).column_letter
        cu = tn.cell(hang_cu + 1, cot_dich)
        moi = tn.cell(hang_moi + 1, cot_dich, f"=SUM({ca}{hang_moi}:{cb}{hang_moi})")
        moi.number_format, moi.font = cu.number_format, Font(bold=True)
        cu.value = None

    n = 0
    for c in range(2, 14):                                    # LỢI NHUẬN B4..M4
        o = ln.cell(4, c)
        if isinstance(o.value, str) and f"!{o.value.split('!')[-1]}".endswith(str(hang_cu)):
            o.value = o.value.replace(str(hang_cu), str(hang_moi))
            n += 1
    if isinstance(cp["Q5"].value, str) and f"P{hang_cu}" in cp["Q5"].value:
        cp["Q5"] = cp["Q5"].value.replace(f"P{hang_cu}", f"P{hang_moi}")
        n += 1
    return n


def _tim_hang_tong(tn) -> int | None:
    """Hàng đang chứa khối tổng của THU NHẬP (ô cột D là =SUM(D3:...))."""
    for r in range(THU_NHAP_HANG_DAU + 1, 200):
        v = tn.cell(r, 4).value
        if isinstance(v, str) and v.upper().startswith(f"=SUM(D{THU_NHAP_HANG_DAU}:"):
            return r
    return None


def lenh_gon_thu_nhap(args) -> None:
    """Kéo khối tổng của THU NHẬP về ngay dưới dòng khách cuối cùng."""
    in_dam("GỌN LẠI SHEET THU NHẬP")
    if excel_dang_mo(XL_SO) and not args.thu:
        sys.exit(f"File đang mở trong Excel: {XL_SO.name} → đóng rồi chạy lại.")
    wb = openpyxl.load_workbook(XL_SO)
    tn, ln, cp = wb["THU NHẬP"], wb["LỢI NHUẬN"], wb["CHI PHÍ"]

    hang_tong = _tim_hang_tong(tn)
    if hang_tong is None:
        sys.exit("Không tìm thấy khối tổng trong THU NHẬP.")
    cuoi = max((r for r in range(THU_NHAP_HANG_DAU, hang_tong) if tn.cell(r, 3).value),
               default=THU_NHAP_HANG_DAU - 1)
    moi = cuoi + 1
    print(f"  khách cuối ở hàng {cuoi} · khối tổng đang ở hàng {hang_tong} → đưa về hàng {moi}")
    if moi == hang_tong:
        print("  Đã gọn sẵn — không làm gì.")
        return
    if args.thu:
        print("\n  [CHẠY KHÔ] Không ghi gì.")
        return

    print(f"  sao lưu → {sao_luu(XL_SO).name}")
    n = _dat_hang_tong(tn, ln, cp, hang_tong, moi)
    luu_wb(wb, XL_SO)
    print(f"  ✓ khối tổng nay ở hàng {moi} (ngay dưới khách cuối), trỏ lại {n} công thức ngoài")


def lenh_soat(_args) -> None:
    in_dam("RÀ SOÁT")
    td = lay_tu_dien()
    tap_ma = set(td.get("ma_det", []))
    van_de = 0

    state = doc_json(STATE_FILE, {})
    moc = state.get("moc_tra_hang")
    if XL_SO.exists():
        ws = openpyxl.load_workbook(XL_SO, data_only=True, read_only=True)[SHEET_SO]
        ngay_cuoi, so_dong, tong = None, 0, 0.0
        for r in ws.iter_rows(min_row=HANG_DAU_SO, min_col=2, max_col=5, values_only=True):
            d = doc_ngay(r[0])
            if d is None:
                continue
            so_dong += 1
            tong += doc_so(r[3]) or 0
            ngay_cuoi = d if ngay_cuoi is None or d > ngay_cuoi else ngay_cuoi
        # Tiền chỉ lên được LỢI NHUẬN nếu mã KH có dòng riêng trong THU NHẬP.
        # So khớp HẠ CHỮ THƯỜNG vì SUMIFS của Excel không phân biệt hoa/thường
        # (sheet Code ghi 'Nhungth' còn THU NHẬP ghi 'nhungth' — vẫn cộng đúng).
        wb_so = openpyxl.load_workbook(XL_SO, read_only=True)
        tn = wb_so["THU NHẬP"]
        co_dong = {str(tn.cell(r, 3).value).strip().lower()
                   for r in range(3, 60) if tn.cell(r, 3).value}
        roi: dict[str, float] = {}
        ws2 = openpyxl.load_workbook(XL_SO, data_only=True, read_only=True)[SHEET_SO]
        for r in ws2.iter_rows(min_row=HANG_DAU_SO, min_col=2, max_col=5, values_only=True):
            if doc_ngay(r[0]) and r[3] and str(r[1] or "").strip().lower() not in co_dong:
                roi[str(r[1]).strip()] = roi.get(str(r[1]).strip(), 0) + (doc_so(r[3]) or 0)
        print(f"  Sổ thật — trả hàng : {so_dong} dòng, tổng {tien(tong)} đ")
        for ma, sotien in roi.items():
            van_de += 1
            print(f"  ⚠ mã KH '{ma}' có {tien(sotien)} đ trong sổ nhưng KHÔNG có dòng trong "
                  f"THU NHẬP → tiền này KHÔNG lên LỢI NHUẬN. Chạy: dh.py them-khach {ma} \"<tên>\"")
        print(f"  Ngày trả hàng cuối : {ngay_cuoi:%d/%m/%Y}" if ngay_cuoi else "  chưa có dòng nào")
        if ngay_cuoi:
            thieu = (dt.date.today() - ngay_cuoi).days
            if thieu > 1:
                van_de += 1
                print(f"  ⚠ {thieu} ngày chưa vào sổ (từ {ngay_cuoi:%d/%m} đến nay) — chạy 'tổng hợp trả hàng'")
    print(f"  Mốc đã đọc Zalo    : {moc or 'chưa có'}")

    if XL_DON.exists():
        wb = openpyxl.load_workbook(XL_DON, data_only=True, read_only=True)
        if SHEET_DON in wb.sheetnames:
            ws = wb[SHEET_DON]
            n, xau = 0, []
            for r in ws.iter_rows(min_row=HANG_DAU_DON, min_col=1, max_col=12, values_only=True):
                if not r[3]:
                    continue
                n += 1
                if not r[5]:
                    xau.append(f"{r[3]} · '{r[4] or ''}' — THIẾU KHỔ, ô Mẫu dệt còn trống")
                elif ma_goc(r[5]) not in tap_ma and "(usb)" not in str(r[5]).lower():
                    xau.append(f"{r[3]} · mã dệt '{r[5]}' không có file thiết kế")
                if not r[11]:
                    xau.append(f"{r[3]} · thiếu Mã KH")
                if not r[2]:
                    xau.append(f"{r[3]} · thiếu tên khách")
            print(f"\n  Sheet đơn hàng     : {n} dòng hàng")
            for x in xau:
                van_de += 1
                print(f"  ⚠ {x}")
        else:
            print(f"\n  ⚠ chưa có sheet '{SHEET_DON}' — chạy: python tools/dh.py sua-file-don")
            van_de += 1

    if ANH_DIR.exists():
        da = set(doc_json(STATE_FILE, {}).get("anh_da_doc", {}))
        anh = [p for p in ANH_DIR.rglob("*") if p.is_file() and p.suffix.lower() in DUOI_ANH]
        chua = sum(1 for p in anh if not da_doc_roi(bam_file(p), da))
        print(f"\n  Ảnh trong Đơn hàng/: {len(anh)} ảnh, {chua} chưa đọc")
        if chua:
            van_de += 1
            print("  ⚠ có ảnh chưa xử lý — chạy 'tổng hợp đơn hàng'")

    print(f"\n  {'✓ không thấy vấn đề' if van_de == 0 else f'⚠ {van_de} việc cần xử lý'}")


# ============================ THÊM KHÁCH MỚI VÀO SỔ THẬT ============================
# Sheet THU NHẬP nguyên bản: khách ở r3..r33 (kín chỗ), ô tổng ở r34 = SUM(x3:x33),
# r35 là 4 ô cộng theo quý. 13 công thức NGOÀI sheet trỏ thẳng vào r34:
#   LỢI NHUẬN!B4..M4 -> 'THU NHẬP'!D34..O34      CHI PHÍ!Q5 -> 'THU NHẬP'!P34
# Nên muốn thêm khách phải dời khối tổng xuống, làm MỘT LẦN, rồi về sau chỉ việc append.

THU_NHAP_HANG_DAU = 3
TN_TONG_CU, TN_QUY_CU = 34, 35
TN_TONG_MOI, TN_QUY_MOI = 60, 61          # chừa r34..r59 = 26 chỗ cho khách mới
THANG_COT = list(range(4, 16))            # D..O = tháng 1..12
COT_TONG_NAM = 16                         # P


def _tong_thu_nhap(p: Path) -> dict[int, float]:
    """Tự cộng doanh thu từng tháng từ TỔNG HỢP — để đối chiếu trước/sau khi mổ."""
    ws = openpyxl.load_workbook(p, data_only=True, read_only=True)[SHEET_SO]
    ra: dict[int, float] = {}
    for r in ws.iter_rows(min_row=HANG_DAU_SO, min_col=2, max_col=5, values_only=True):
        d, st = doc_ngay(r[0]), doc_so(r[3])
        if d and st:
            ra[d.month] = ra.get(d.month, 0.0) + st
    return ra


def lenh_mo_rong_thu_nhap(_args) -> None:
    """MỘT LẦN: dời khối tổng của THU NHẬP xuống r60 và nới vùng cộng ra r3:r59."""
    in_dam("MỞ RỘNG SHEET THU NHẬP (chạy một lần)")
    if excel_dang_mo(XL_SO):
        sys.exit(f"File đang mở trong Excel: {XL_SO.name} → đóng rồi chạy lại.")
    wb = openpyxl.load_workbook(XL_SO)
    tn, ln, cp = wb["THU NHẬP"], wb["LỢI NHUẬN"], wb["CHI PHÍ"]

    # Dấu hiệu "đã mở rộng" phải là SỰ CÓ MẶT của khối tổng mới ở r60 — không được dùng
    # "D34 rỗng", vì r34 sẽ bị khách mới đầu tiên chiếm chỗ ngay sau khi mở rộng.
    if tn.cell(TN_TONG_MOI, 4).value not in (None, ""):
        print(f"  Đã mở rộng trước đó (khối tổng đang ở r{TN_TONG_MOI}) — không làm gì.")
        return

    truoc = _tong_thu_nhap(XL_SO)
    print(f"  doanh thu tự cộng từ TỔNG HỢP: {tien(sum(truoc.values()))} đ")
    print(f"  sao lưu → {sao_luu(XL_SO).name}")

    # 1. dựng khối tổng mới ở r60, vùng cộng r3:r59 (không chứa chính nó)
    for c in THANG_COT + [COT_TONG_NAM]:
        cu = tn.cell(TN_TONG_CU, c)
        col = cu.column_letter
        moi = tn.cell(TN_TONG_MOI, c, f"=SUM({col}{THU_NHAP_HANG_DAU}:{col}{TN_TONG_MOI - 1})")
        moi.number_format, moi.font = cu.number_format, Font(bold=True)
    tn.cell(TN_TONG_MOI, 2, "TỔNG").font = Font(bold=True)

    # 2. dời 4 ô cộng theo quý xuống r61 (bám theo hàng tổng mới)
    for cot_dich, (a, b) in ((6, (4, 6)), (9, (7, 9)), (12, (10, 12)), (15, (13, 15))):
        ca = tn.cell(TN_TONG_MOI, a).column_letter
        cb = tn.cell(TN_TONG_MOI, b).column_letter
        cu = tn.cell(TN_QUY_CU, cot_dich)
        moi = tn.cell(TN_QUY_MOI, cot_dich, f"=SUM({ca}{TN_TONG_MOI}:{cb}{TN_TONG_MOI})")
        moi.number_format, moi.font = cu.number_format, Font(bold=True)

    # 3. xoá khối cũ (nếu để lại thì bị cộng hai lần vào vùng r3:r59)
    for c in range(1, 17):
        tn.cell(TN_TONG_CU, c).value = None
        tn.cell(TN_QUY_CU, c).value = None
    print(f"  dời khối tổng r{TN_TONG_CU} → r{TN_TONG_MOI}, vùng cộng nới thành "
          f"r{THU_NHAP_HANG_DAU}:r{TN_TONG_MOI - 1} (chừa {TN_TONG_MOI - TN_TONG_CU} chỗ cho khách mới)")

    # 4. trỏ lại 13 công thức bên ngoài
    n = 0
    for c in range(2, 14):                                   # LỢI NHUẬN B4..M4
        o = ln.cell(4, c)
        if isinstance(o.value, str) and f"!{o.value.split('!')[-1]}" and str(TN_TONG_CU) in str(o.value):
            o.value = o.value.replace(f"{TN_TONG_CU}", f"{TN_TONG_MOI}")
            n += 1
    if isinstance(cp["Q5"].value, str) and str(TN_TONG_CU) in cp["Q5"].value:
        cp["Q5"] = cp["Q5"].value.replace(f"P{TN_TONG_CU}", f"P{TN_TONG_MOI}")
        n += 1
    print(f"  trỏ lại {n} công thức bên ngoài (LỢI NHUẬN + CHI PHÍ)")

    luu_wb(wb, XL_SO)
    print(f"\n  ✓ đã lưu. Doanh thu kỳ vọng không đổi: {tien(sum(truoc.values()))} đ")
    print("  Anh Đức mở file kiểm sheet LỢI NHUẬN — các tháng phải y như trước.")


def lenh_them_khach(args) -> None:
    """Thêm một khách mới vào sheet Code và sheet THU NHẬP của sổ thật."""
    in_dam("THÊM KHÁCH MỚI VÀO SỔ THẬT")
    if excel_dang_mo(XL_SO):
        sys.exit(f"File đang mở trong Excel: {XL_SO.name} → đóng rồi chạy lại.")
    ma, ten = args.ma_kh.strip(), args.ten.strip()
    # Sổ của anh Đức có sẵn mã viết hoa chữ đầu (ví dụ 'Nhungth') nên phải cho phép.
    if not re.fullmatch(r"[A-Za-z0-9]+", ma):
        sys.exit(f"Mã KH '{ma}' phải là chữ/số liền nhau, không dấu, không khoảng trắng (ví dụ: bform).")

    wb = openpyxl.load_workbook(XL_SO)
    code, tn = wb["Code"], wb["THU NHẬP"]

    # --- sheet Code: khối Mã KH ---
    hang_kh = [r for r in range(15, 60) if code.cell(r, 2).value]
    da_co = {str(code.cell(r, 2).value).strip(): r for r in hang_kh}
    if ma in da_co:
        print(f"  mã '{ma}' đã có ở Code!B{da_co[ma]} = {code.cell(da_co[ma], 3).value!r} — bỏ qua sheet Code")
        hang_moi_code = None
    else:
        hang_moi_code = max(hang_kh) + 1
        if code.cell(hang_moi_code, 2).value or khong_dau(code.cell(hang_moi_code + 1, 2).value or "").startswith("ma "):
            code.insert_rows(hang_moi_code)
        stt = max((int(code.cell(r, 1).value) for r in hang_kh
                   if isinstance(code.cell(r, 1).value, (int, float))), default=0) + 1
        code.cell(hang_moi_code, 1, stt)
        code.cell(hang_moi_code, 2, ma)
        code.cell(hang_moi_code, 3, ten)
        if args.dia_chi:
            code.cell(hang_moi_code, 5, args.dia_chi)
        if args.mat_hang:
            code.cell(hang_moi_code, 6, args.mat_hang)

    # --- sheet THU NHẬP ---
    # Khách mới chiếm đúng hàng đang chứa khối tổng, rồi đẩy khối tổng xuống 1 hàng.
    # Nhờ vậy bảng luôn gọn: tổng nằm ngay dưới khách cuối, không có hàng trống ở giữa.
    ln, cp = wb["LỢI NHUẬN"], wb["CHI PHÍ"]
    hang_tong = _tim_hang_tong(tn)
    if hang_tong is None:
        sys.exit("Không tìm thấy khối tổng trong THU NHẬP — kiểm tra lại file.")
    co_ma = {str(tn.cell(r, 3).value).strip().lower() for r in range(THU_NHAP_HANG_DAU, hang_tong)
             if tn.cell(r, 3).value}
    if ma.lower() in co_ma:
        print(f"  mã '{ma}' đã có trong THU NHẬP — bỏ qua")
        hang_moi_tn = None
    else:
        hang_moi_tn = hang_tong
        print(f"  sao lưu → {sao_luu(XL_SO).name}")
        n_ref = _dat_hang_tong(tn, ln, cp, hang_tong, hang_tong + 1)
        print(f"  đẩy khối tổng r{hang_tong} → r{hang_tong + 1}, trỏ lại {n_ref} công thức ngoài")
        stt = max((int(tn.cell(r, 1).value) for r in range(THU_NHAP_HANG_DAU, hang_tong)
                   if isinstance(tn.cell(r, 1).value, (int, float))), default=0) + 1
        tn.cell(hang_moi_tn, 1, stt)
        tn.cell(hang_moi_tn, 2, ten)
        tn.cell(hang_moi_tn, 3, ma)
        # Dùng dạng SUMIFS phẳng (giống các tháng 5–12 trong file gốc), không dùng biến thể
        # lookup() ở tháng 1–4 — dạng phẳng ngắn và đúng vì cột A của TỔNG HỢP là =MONTH(B).
        for c in THANG_COT:
            col = tn.cell(hang_moi_tn, c).column_letter
            o = tn.cell(hang_moi_tn, c,
                        f"=SUMIFS('{SHEET_SO}'!$E:$E,'{SHEET_SO}'!$C:$C,$C{hang_moi_tn},"
                        f"'{SHEET_SO}'!$A:$A,{col}$2)")
            o.number_format = tn.cell(THU_NHAP_HANG_DAU, c).number_format
        o = tn.cell(hang_moi_tn, COT_TONG_NAM, f"=SUM(D{hang_moi_tn}:O{hang_moi_tn})")
        o.number_format = tn.cell(THU_NHAP_HANG_DAU, COT_TONG_NAM).number_format

    if hang_moi_code is None and hang_moi_tn is None:
        print("\n  Không có gì thay đổi.")
        return
    luu_wb(wb, XL_SO)
    if hang_moi_code:
        print(f"  ✓ Code    : thêm '{ma}' = {ten!r} ở hàng {hang_moi_code}")
    if hang_moi_tn:
        print(f"  ✓ THU NHẬP: thêm '{ma}' ở hàng {hang_moi_tn}, đủ 12 công thức SUMIFS + tổng năm")
    print("\n  Chạy tiếp: python tools/dh.py xay-tu-dien   (để từ điển nhận tên khách này)")


# ============================ SỬA FILE ĐƠN HÀNG ============================


def lenh_sua_file_don(args) -> None:
    """Chuẩn hoá Tong-hop-don-hang-Xbo-Duc-Lan-2026.xlsx cho cả năm."""
    in_dam("CHUẨN HOÁ FILE EXCEL ĐƠN HÀNG")
    if not XL_DON.exists():
        sys.exit(f"Không thấy {XL_DON}")
    if excel_dang_mo(XL_DON):
        sys.exit(f"File đang mở trong Excel: {XL_DON.name} → đóng rồi chạy lại.")

    wb = openpyxl.load_workbook(XL_DON)
    print(f"  sheet hiện có: {wb.sheetnames}")
    if SHEET_DON in wb.sheetnames:
        ws = wb[SHEET_DON]
        print(f"  '{SHEET_DON}' đã tồn tại — chỉ kiểm lại khối tổng và cột Mã KH")
    else:
        ws = wb[wb.sheetnames[0]]
        cu = ws.title
        ws.title = SHEET_DON
        print(f"  đổi tên sheet '{cu}' → '{SHEET_DON}'")

    print(f"  sao lưu → {sao_luu(XL_DON).name}")

    # bỏ khối tổng lỗi #REF! đang nằm giữa vùng dữ liệu (H6:J9)
    for r in range(6, 10):
        for c in range(8, 11):
            ws.cell(r, c).value = None
    for m in [str(x) for x in ws.merged_cells.ranges]:
        if m not in ("A1:K1",):
            try:
                ws.unmerge_cells(m)
            except (KeyError, ValueError):
                pass
    print("  đã xoá khối tổng lỗi #REF! ở H6:J9 (nằm chắn vùng dữ liệu)")

    # thêm cột Mã KH
    dam = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    xanh = PatternFill("solid", fgColor="1A56DB")
    vien = Border(*[Side(style="thin", color="D0D7DE")] * 4)
    o = ws.cell(HANG_HEADER_DON, 12, "Mã KH")
    o.font, o.fill, o.border = dam, xanh, vien
    o.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["L"].width = 12
    print("  đã thêm cột L 'Mã KH'")

    # dựng lại khối tổng ở N:P, ngoài vùng dữ liệu
    ws["N4"] = "TỔNG THEO ĐƠN VỊ"
    ws["N4"].font = Font(name="Calibri", size=11, bold=True)
    for i, dv in enumerate(("bộ", "cái", "cổ", "kg"), start=5):
        ws.cell(i, 14, f"Tổng ({dv})")
        c = ws.cell(i, 15, f'=SUMIF($J$5:$J$5000,"{dv}",$I$5:$I$5000)')
        c.number_format = "#,##0"
        ws.cell(i, 16, dv)
    ws["N10"] = "Số đơn"
    ws["O10"] = "=SUMPRODUCT((($D$5:$D$5000<>\"\")/COUNTIF($D$5:$D$5000,$D$5:$D$5000&\"\")))"
    ws["N11"] = "Số dòng hàng"
    ws["O11"] = '=COUNTA($D$5:$D$5000)'
    for col, w in (("N", 18), ("O", 12), ("P", 7)):
        ws.column_dimensions[col].width = w
    print("  đã dựng khối tổng mới ở N4:P11 (SUMIF theo cột Đơn vị, ngoài vùng dữ liệu)")

    ws.freeze_panes = "A5"
    luu_wb(wb, XL_DON)
    print(f"\n  ✓ đã lưu {XL_DON.name}")


# ============================ CLI ============================


def main() -> None:
    p = argparse.ArgumentParser(prog="dh.py", description="Công cụ đơn hàng Xưởng Bo Đức Lan")
    sub = p.add_subparsers(dest="lenh", required=True)

    sub.add_parser("xay-tu-dien", help="dựng lại tools/tu-dien.json").set_defaults(f=lenh_xay_tu_dien)

    a = sub.add_parser("anh-moi", help="liệt kê + thu nhỏ ảnh chưa xử lý")
    a.add_argument("--canh", type=int, default=1400, help="cạnh dài tối đa (mặc định 1400px)")
    a.add_argument("--ra", help="thư mục chứa ảnh thu nhỏ")
    a.set_defaults(f=lenh_anh_moi)

    for ten, fn, mo in (("don-them", lenh_don_them, "ghi đơn hàng vào Excel đơn hàng"),
                        ("tra-them", lenh_tra_them, "ghi trả hàng vào sổ thật")):
        s = sub.add_parser(ten, help=mo)
        s.add_argument("--json", required=True, help="file JSON do LLM trích ra")
        s.add_argument("--thu", action="store_true", help="chạy khô, không ghi gì")
        s.set_defaults(f=fn)

    al = sub.add_parser("them-alias", help="khai báo tên khách mới cho từ điển")
    al.add_argument("ten", help='tên như anh Đức viết, ví dụ "Mai Liên Time"')
    al.add_argument("ma_kh", help="mã KH trong sheet Code, ví dụ halien")
    al.add_argument("--ep", action="store_true", help="ghi kể cả khi mã chưa có trong sổ")
    al.set_defaults(f=lenh_them_alias)

    sk = sub.add_parser("sua-kho", help="điền khổ cho đơn bị THIẾU KHỔ, dựng lại mã dệt")
    sk.add_argument("ma_don", help="mã đơn, ví dụ Ms62")
    sk.add_argument("ho", help="họ mã: Co / Ta / Ga / GDO ...")
    sk.add_argument("kho", help="khổ, ví dụ 42x75")
    sk.add_argument("--thu", action="store_true", help="chạy khô, không ghi gì")
    sk.set_defaults(f=lenh_sua_kho)

    xt = sub.add_parser("xoa-tra", help="xoá dòng trong khối TRẢ HÀNG (khi lỡ ghi trùng)")
    xt.add_argument("hang", nargs="+", help="số hàng cần xoá, ví dụ 237 238")
    xt.add_argument("--thu", action="store_true", help="chạy khô, không xoá gì")
    xt.set_defaults(f=lenh_xoa_tra)

    gt = sub.add_parser("gon-thu-nhap", help="kéo khối tổng THU NHẬP về sát dòng khách cuối")
    gt.add_argument("--thu", action="store_true", help="chạy khô, không ghi gì")
    gt.set_defaults(f=lenh_gon_thu_nhap)

    dt_ = sub.add_parser("don-tra", help="dồn khối TRẢ HÀNG, lấp hàng trống ở giữa")
    dt_.add_argument("--thu", action="store_true", help="chạy khô, không ghi gì")
    dt_.set_defaults(f=lenh_don_tra)

    ch = sub.add_parser("chuan-hoa-cong-thuc",
                        help="sửa =MONTH() sinh số 1 rác và bỏ lookup() trong SUMIFS")
    ch.add_argument("--thu", action="store_true", help="chạy khô, không ghi gì")
    ch.set_defaults(f=lenh_chuan_hoa_cong_thuc)

    db = sub.add_parser("dong-bo-thu-nhap", help="đánh lại STT + đồng bộ định dạng sheet THU NHẬP")
    db.add_argument("--mau", help="số hàng dùng làm mẫu định dạng")
    db.add_argument("--thu", action="store_true", help="chạy khô, không ghi gì")
    db.set_defaults(f=lenh_dong_bo_thu_nhap)

    sub.add_parser("soat", help="rà soát sức khoẻ dữ liệu").set_defaults(f=lenh_soat)
    sub.add_parser("sua-file-don", help="chuẩn hoá file Excel đơn hàng").set_defaults(f=lenh_sua_file_don)
    sub.add_parser("mo-rong-thu-nhap",
                   help="MỘT LẦN: dời khối tổng sheet THU NHẬP để thêm được khách mới"
                   ).set_defaults(f=lenh_mo_rong_thu_nhap)

    tk = sub.add_parser("them-khach", help="thêm khách mới vào sheet Code + THU NHẬP của sổ thật")
    tk.add_argument("ma_kh", help="mã KH chữ thường không dấu, ví dụ bform")
    tk.add_argument("ten", help='tên khách, ví dụ "Cty Bform"')
    tk.add_argument("--dia-chi", default="", help="địa chỉ (để sinh alias viết tắt)")
    tk.add_argument("--mat-hang", default="bo", help="SP/Dịch vụ, mặc định 'bo'")
    tk.set_defaults(f=lenh_them_khach)

    m = sub.add_parser("ma-det", help="thử quy tắc đuôi mã dệt")
    m.add_argument("ho")
    m.add_argument("kho")
    m.add_argument("kieu")
    m.set_defaults(f=lambda ar: print(" -> ".join(dung_ma_det(ar.ho, ar.kho, ar.kieu))))

    args = p.parse_args()
    args.f(args)


if __name__ == "__main__":
    main()
